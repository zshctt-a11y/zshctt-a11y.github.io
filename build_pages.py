# -*- coding: utf-8 -*-
"""Generate per-product detail pages from quotation.xlsx (Catalogue + Specifications)."""
import openpyxl, re, os, html, json

wb = openpyxl.load_workbook('quotation.xlsx', read_only=True)

def slug(n):
    if n == "SMD-Z1 Pro+": return "smd-z1-pro-plus"
    return re.sub(r'[^a-z0-9]+', '-', n.lower()).strip('-')

# --- prices + headline specs from Catalogue ---
cat = wb['Catalogue']
prices, headline = {}, {}
name_rows = []
for row in cat.iter_rows(min_row=1, max_row=63, max_col=8):
    for c in row:
        if c.value and isinstance(c.value, str) and 2 <= len(c.value.strip()) <= 20:
            pass
# name rows are those whose following row contains specs and next-next contains price
vals = {}
for ri, row in enumerate(cat.iter_rows(min_row=1, max_row=63, max_col=8, values_only=True), start=1):
    for ci, v in enumerate(row, start=1):
        vals[(ri, ci)] = v
for (r, col), v in list(vals.items()):
    if not v or not isinstance(v, str): continue
    spec = vals.get((r+1, col)); price = vals.get((r+2, col))
    if spec and price is not None:
        try:
            p = float(str(price).replace(',', ''))
        except ValueError:
            continue
        name = v.strip()
        parts = [x.strip() for x in str(spec).split('|')]
        prices[name] = p
        headline[name] = parts

# --- full specs from Specifications ---
spec_ws = wb['Specifications']
configs = {}
HDR = ["Configuration","Motor","Battery","Controller","Max speed","Range","Brakes / Suspension","Tyre","Charging","Gross weight","Package L*W*H (cm)","Container loading","Certification"]
cur_line = None
for row in spec_ws.iter_rows(min_row=1, max_row=65, max_col=16, values_only=True):
    first = str(row[0] or '')
    if first.startswith('LINE'):
        cur_line = first; continue
    if first in ('A','B','C') and row[2]:
        name = str(row[2]).strip()
        cfg = {}
        for i, h in enumerate(HDR):
            v = row[3+i] if 3+i < len(row) else None
            cfg[h] = str(v).strip().replace('\n', ' / ') if v not in (None, '', '-') else ''
        cfg['_line'] = first
        configs.setdefault(name, []).append(cfg)

# merge: every priced model
models = []
for name, p in prices.items():
    cfgs = configs.get(name, [])
    line = cfgs[0]['_line'] if cfgs else ('C' if 'Bike' in name or 'Moped' in name or 'Tricycle' in name or name.startswith(('L26','ZM22','DK400')) else 'A')
    models.append(dict(name=name, slug=slug(name), price=p, line=line,
                       headline=headline.get(name, []), configs=cfgs))

LINE_NAMES = {'A': 'E-Series · Performance', 'B': 'SMD · Factory Direct', 'C': 'E-Bikes & Mopeds'}

TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{name} — MG MAX | TUFRONT</title>
<style>
:root{{--red:#e8102e;--red-dark:#a80c22;--black:#0a0a0b;--panel:#131315;--silver:#c9ccd1;--white:#f5f6f7;--muted:#8a8d93;--line:#232326}}
*{{margin:0;padding:0;box-sizing:border-box}}
body{{background:var(--black);color:var(--white);font-family:"Segoe UI",-apple-system,"Helvetica Neue",Arial,sans-serif}}
h1,h2,h3{{font-family:"Arial Black","Segoe UI",sans-serif;font-style:italic;letter-spacing:-.02em}}
a{{color:inherit;text-decoration:none}}
.wrap{{max-width:1200px;margin:0 auto;padding:0 28px}}
nav{{position:sticky;top:0;z-index:100;background:rgba(10,10,11,.85);backdrop-filter:blur(14px);border-bottom:1px solid var(--line)}}
.nav-inner{{display:flex;align-items:center;justify-content:space-between;height:68px;max-width:1200px;margin:0 auto;padding:0 28px}}
.brand{{display:flex;align-items:center;gap:12px}}
.brand img{{height:40px;width:40px;border-radius:50%}}
.brand b{{font-size:19px;font-style:italic}}.brand b span{{color:var(--red)}}
.back{{font-size:13px;color:var(--silver);text-transform:uppercase;letter-spacing:.12em}}
.back:hover{{color:var(--red)}}
.hero{{display:grid;grid-template-columns:1fr 1fr;gap:48px;align-items:center;padding:64px 0 40px}}
.kicker{{display:inline-flex;align-items:center;gap:10px;color:var(--red);font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.3em;margin-bottom:16px}}
.kicker::before{{content:"";width:34px;height:2px;background:var(--red)}}
h1{{font-size:clamp(40px,5.5vw,72px);text-transform:uppercase;line-height:.95}}
.lead{{color:var(--silver);margin:20px 0 26px;line-height:1.75;font-size:16px}}
.price-box{{display:flex;align-items:baseline;gap:14px;margin-bottom:28px}}
.price-box b{{font-size:44px;font-style:italic;color:var(--red)}}
.price-box span{{color:var(--muted);font-size:12px;text-transform:uppercase;letter-spacing:.14em}}
.ctas{{display:flex;gap:14px;flex-wrap:wrap}}
.btn{{display:inline-block;padding:15px 32px;border-radius:4px;font-weight:800;font-size:13px;text-transform:uppercase;letter-spacing:.12em;transition:.2s}}
.btn-red{{background:var(--red);color:#fff;clip-path:polygon(0 0,100% 0,94% 100%,0 100%)}}
.btn-red:hover{{background:var(--red-dark)}}
.btn-ghost{{border:1px solid var(--silver)}}
.btn-ghost:hover{{border-color:var(--red);color:var(--red)}}
.imgbox{{background:radial-gradient(360px 280px at 50% 45%,rgba(232,16,46,.14),transparent 70%),var(--panel);border:1px solid var(--line);border-radius:14px;display:flex;align-items:center;justify-content:center;padding:34px}}
.imgbox img{{max-width:100%;max-height:460px;object-fit:contain;filter:drop-shadow(0 24px 44px rgba(0,0,0,.6))}}
.stats{{display:grid;grid-template-columns:repeat(3,1fr);gap:1px;background:var(--line);border:1px solid var(--line);margin:14px 0 60px}}
.stat{{background:var(--black);padding:22px 18px;text-align:center}}
.stat b{{display:block;font-size:22px;font-style:italic;color:var(--white)}}
.stat span{{font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:.14em}}
section{{padding:20px 0 70px}}
h2{{font-size:clamp(26px,3vw,38px);text-transform:uppercase;margin-bottom:26px}}
table{{width:100%;border-collapse:collapse;font-size:14px}}
th,td{{padding:13px 14px;border:1px solid var(--line);text-align:left;vertical-align:top}}
th{{background:var(--panel);color:var(--red);text-transform:uppercase;font-size:11px;letter-spacing:.14em;white-space:nowrap}}
td:first-child{{font-weight:700;white-space:nowrap}}
tr:nth-child(even) td{{background:#101012}}
.note{{color:var(--muted);font-size:12px;margin-top:14px}}
.cert{{display:inline-block;background:var(--panel);border:1px solid var(--line);border-radius:4px;padding:8px 14px;margin:4px 6px 0 0;font-size:12px;color:var(--silver)}}
footer{{border-top:1px solid var(--line);padding:34px 0;color:var(--muted);font-size:13px;text-align:center}}
@media(max-width:900px){{.hero{{grid-template-columns:1fr}}.stats{{grid-template-columns:repeat(2,1fr)}}}}
</style>
</head>
<body>
<nav><div class="nav-inner">
  <a class="brand" href="../index.html"><img src="../logo.jpg" alt="MG MAX"><b>MG <span>MAX</span></b></a>
  <a class="back" href="../index.html#scooters">← Back to Lineup</a>
</div></nav>
<div class="wrap">
  <div class="hero">
    <div>
      <div class="kicker">{line_name}</div>
      <h1>{name}</h1>
      <p class="lead">{lead}</p>
      <div class="price-box"><b>${price}</b><span>Retail · DDP Door · USD</span></div>
      <div class="ctas">
        <a class="btn btn-red" href="mailto:sales@tufront.com?subject=Quote%20request%3A%20{qname}">Request Quote</a>
        <a class="btn btn-ghost" href="../index.html#scooters">View All Models</a>
      </div>
    </div>
    <div class="imgbox"><img src="img/{slug}.jpg" alt="{name}"></div>
  </div>
  <div class="stats">{stats}</div>
  <section>
    <h2>Technical <span style="color:var(--red)">Specifications</span></h2>
    <div style="overflow-x:auto"><table>
      <tr><th>Spec</th>{cfg_headers}</tr>
      {spec_rows}
    </table></div>
    <p class="note">All scooters ship EU-compliant at 25 km/h; documented 5× M-button procedure unlocks full speed for closed-course use. Max load 150 kg on all scooter models.</p>
  </section>
  <section>
    <h2>Compliance &amp; <span style="color:var(--red)">Certification</span></h2>
    {certs}
    <p class="note">12-month warranty on frame, motor, controller and battery pack · spare-parts support for 3 years.</p>
  </section>
</div>
<footer>© 2026 TUFRONT TECHNOLOGY CO., LTD. · sales@tufront.com · Alibaba.com store: tufronthk</footer>
</body></html>"""

LEADS = {
 'A': "Own-brand E-Series performance scooter — factory tooling and assembly controlled by TUFRONT, with motor thermal protection and sealed potted battery packs as standard.",
 'B': "Factory-direct SMD model with multiple battery capacity options — tune price point and range per market from a single SKU.",
 'C': "European-warehouse e-mobility model — DDP door pricing, dropship a single unit or a full pallet, 3–7 day delivery.",
}
DEFAULT_CERTS = {
 'A': ["CE-EN17128","CE-MD","CE-RED","CE-ROHS","UL2272","FCC ID","UL2271 Battery","IEC 62133","UN38.3","EU Battery Regulation"],
 'B': ["CE","ROHS","UN38.3 Battery","MSDS","UL2272 on request"],
 'C': ["CE","ROHS","EN15194 (pedal-assist)","UN38.3 Battery","MSDS"],
}

os.makedirs('products', exist_ok=True)
index = []
for m in models:
    cfgs = m['configs'] or [{}]
    # stat strip: first non-empty of motor/speed/range across configs + price
    def pick(key):
        for c in cfgs:
            if c.get(key): return c[key]
        return ''
    motor, speed, rng = pick('Motor'), pick('Max speed'), pick('Range')
    stats = ""
    for lab, v in [("Motor", motor), ("Max Speed", speed), ("Range", rng)]:
        stats += f'<div class="stat"><b>{html.escape(v) or "—"}</b><span>{lab}</span></div>'
    # spec table: rows = fields, columns = configurations
    keys = ["Configuration","Motor","Battery","Max speed","Range","Brakes / Suspension","Tyre","Charging","Gross weight","Package L*W*H (cm)","Container loading"]
    cfg_headers = "".join(f"<th>Config {i+1}</th>" for i in range(len(cfgs)))
    rows = ""
    for k in keys:
        tds = "".join(f"<td>{html.escape(c.get(k,'') or '—')}</td>" for c in cfgs)
        rows += f"<tr><td>{k}</td>{tds}</tr>"
    certs = "".join(f'<span class="cert">{c}</span>' for c in DEFAULT_CERTS[m['line']])
    page = TEMPLATE.format(
        name=html.escape(m['name']), qname=m['name'].replace(' ', '%20'),
        slug=m['slug'], line_name=LINE_NAMES[m['line']],
        lead=LEADS[m['line']], price=f"{m['price']:,.2f}",
        stats=stats, cfg_headers=cfg_headers, spec_rows=rows, certs=certs)
    with open(f"products/{m['slug']}.html", 'w', encoding='utf-8') as f:
        f.write(page)
    index.append(dict(name=m['name'], slug=m['slug'], line=m['line'], price=m['price'],
                      motor=motor, speed=speed, rng=rng,
                      img=f"products/img/{m['slug']}.jpg", url=f"products/{m['slug']}.html",
                      tag=m['headline'][3] if len(m['headline']) > 3 else LINE_NAMES[m['line']]))

json.dump(index, open('products/index.json', 'w'), ensure_ascii=False, indent=1)
print(f"generated {len(index)} product pages")
